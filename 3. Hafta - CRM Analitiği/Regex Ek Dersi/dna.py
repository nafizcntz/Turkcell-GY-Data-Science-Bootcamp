import openpyxl
import re
from pathlib import Path

regex = r"(\w)"

xlsx_file = Path('egfr_snps.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active

for i,row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=1):
    sheet.cell(row=i+1, column=2).value = row[1].value.replace(",", "")
    if row[2].value is None:
        continue
    matches = re.finditer(regex, row[2].value, re.IGNORECASE)

    new_col = ""

    for matchNum, match in enumerate(matches, start=1):

        for groupNum in range(0, len(match.groups())):
            groupNum = groupNum + 1
            sheet.cell(row=i+1, column=matchNum+4).value = match.group(groupNum)
            new_col += match.group(groupNum)
    sheet.cell(row=i+1, column=4).value = new_col


wb_obj.save(r"./new.xlsx")